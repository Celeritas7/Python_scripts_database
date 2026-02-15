"""
Extract Text Data from ALL Excel Files (Remove Images & Embedded Objects)
=========================================================================
Usage:
    pip install openpyxl
    python extract_excel_text.py

Place this script in the folder containing your Excel files, or run it
from that folder. It will process ALL .xlsx/.xls files found and save
cleaned versions in a 'cleaned' subfolder.
"""

from openpyxl import load_workbook
from openpyxl import Workbook
import os
import glob


def get_base_dir():
    """
    Smart directory detection:
    - If called from a different folder (e.g. via tools, IDE, or another script),
      use the current working directory.
    - If run directly from its own folder, use the script's directory.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    work_dir = os.getcwd()

    if work_dir != script_dir:
        base_dir = work_dir
    else:
        base_dir = script_dir

    return base_dir


def extract_text_only(input_path, output_path):
    print(f"\n  Loading: {os.path.basename(input_path)} ...")
    wb_source = load_workbook(input_path, data_only=True)
    wb_new = Workbook()
    wb_new.remove(wb_new.active)

    for sheet_name in wb_source.sheetnames:
        ws_source = wb_source[sheet_name]
        ws_new = wb_new.create_sheet(title=sheet_name)

        print(f"    Sheet: '{sheet_name}' ({ws_source.max_row} rows x {ws_source.max_column} cols)")

        for row in ws_source.iter_rows(min_row=1, max_row=ws_source.max_row,
                                        min_col=1, max_col=ws_source.max_column):
            for cell in row:
                new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)

                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.alignment = cell.alignment.copy()
                    new_cell.number_format = cell.number_format

        for col_letter, dim in ws_source.column_dimensions.items():
            ws_new.column_dimensions[col_letter].width = dim.width

        for merged_range in ws_source.merged_cells.ranges:
            ws_new.merge_cells(str(merged_range))

    wb_new.save(output_path)
    input_size = os.path.getsize(input_path) / (1024 * 1024)
    output_size = os.path.getsize(output_path) / (1024 * 1024)
    print(f"    {input_size:.1f} MB --> {output_size:.1f} MB  ✓")


if __name__ == "__main__":
    base_dir = get_base_dir()
    print(f"Working directory: {base_dir}")

    # Find all Excel files in the folder
    excel_files = sorted(
        glob.glob(os.path.join(base_dir, "*.xlsx")) +
        glob.glob(os.path.join(base_dir, "*.xls"))
    )

    if not excel_files:
        print("\nNo Excel files (.xlsx / .xls) found in this folder.")
        print("Please run this script from the folder containing your Excel files.")
        exit(1)

    print(f"Found {len(excel_files)} Excel file(s)")

    # Create 'cleaned' subfolder for output
    output_dir = os.path.join(base_dir, "cleaned")
    os.makedirs(output_dir, exist_ok=True)

    success = 0
    failed = 0

    for filepath in excel_files:
        filename = os.path.basename(filepath)
        output_path = os.path.join(output_dir, filename)

        try:
            extract_text_only(filepath, output_path)
            success += 1
        except Exception as e:
            print(f"\n  ✗ Failed: {filename} — {e}")
            failed += 1

    print(f"\n{'='*50}")
    print(f"Complete! {success} processed, {failed} failed.")
    print(f"Cleaned files saved to: {output_dir}")

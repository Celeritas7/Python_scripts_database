"""
Clean Excel Work Instructions — Remove Images, Keep Data Inside Excel
======================================================================
Usage:
    pip install openpyxl
    python extract_assembly_data.py

For each Excel file, this script:
  1. Extracts assembly steps, parts, and comments from the drawing layer
  2. Removes all embedded images (keeps text shapes, annotations)
  3. Writes extracted data into a new "Extracted Data" sheet inside the same Excel
  4. Saves cleaned file to _extracted/ folder

No JSON needed — everything stays inside the Excel file.
"""

import zipfile
import xml.etree.ElementTree as ET
import re
import os
import glob
import io
import shutil
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ─── Configuration ───────────────────────────────────────
OUTPUT_FOLDER_NAME = "_extracted"
EXTRACTED_SHEET_NAME = "Extracted Data"

VIEW_LABELS = {"top view", "bottom view", "front view", "back view",
               "side view", "left view", "right view", "isometric view",
               "top", "bottom", "front", "back", "left", "right"}

PART_NUMBER_PATTERN = re.compile(
    r'[A-Z]{2,5}-\d{5,10}-\d+-\d+'
    r'|[A-Z]{2,5}-\d{3,}'
    r'|\d{3,}-\d{3,}'
)

STEP_PATTERN = re.compile(r'^(\d+)\)\s*')
# ─────────────────────────────────────────────────────────

NS_XDR = '{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}'
NS_A = '{http://schemas.openxmlformats.org/drawingml/2006/main}'
NS_R = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'


def get_base_dir():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    work_dir = os.getcwd()
    if work_dir != script_dir:
        base_dir = work_dir
    else:
        base_dir = script_dir
    return base_dir


# ─── Drawing Text Extraction ────────────────────────────

def get_full_text(elem):
    paragraphs = []
    for p in elem.findall(f'.//{NS_A}p'):
        runs = []
        for r in p.findall(f'{NS_A}r'):
            t = r.find(f'{NS_A}t')
            if t is not None and t.text:
                runs.append(t.text)
        if runs:
            paragraphs.append("".join(runs))
    return "\n".join(paragraphs)


def extract_all_texts(elem):
    texts = []
    tag = elem.tag.split('}')[-1]
    if tag == 'sp':
        text = get_full_text(elem).strip()
        if text:
            texts.append(text)
    elif tag in ('grpSp', 'oneCellAnchor', 'twoCellAnchor'):
        for child in elem:
            texts.extend(extract_all_texts(child))
    return texts


def extract_texts_from_zip(zip_file):
    all_files = zip_file.namelist()
    drawing_files = [f for f in all_files
                     if f.startswith('xl/drawings/') and f.endswith('.xml') and '_rels' not in f]
    all_texts = []
    for df in drawing_files:
        drawing_xml = zip_file.read(df).decode("utf-8")
        root = ET.fromstring(drawing_xml)
        for anchor in root:
            if 'Anchor' not in anchor.tag.split('}')[-1]:
                continue
            all_texts.extend(extract_all_texts(anchor))
    return all_texts


# ─── Comment Extraction ─────────────────────────────────

def extract_all_comments(zip_file):
    comments = []
    all_files = zip_file.namelist()

    for cf in [f for f in all_files if f.startswith('xl/comments') and f.endswith('.xml')]:
        xml_bytes = zip_file.read(cf)
        root = ET.fromstring(xml_bytes)
        ns = root.tag.split('}')[0] + '}' if root.tag.startswith('{') else ''
        authors_elem = root.find(f'{ns}authors')
        authors = [a.text or "" for a in authors_elem.findall(f'{ns}author')] if authors_elem is not None else []
        comment_list = root.find(f'{ns}commentList')
        if comment_list is not None:
            for c in comment_list.findall(f'{ns}comment'):
                cell_ref = c.get('ref', '?')
                author_id = int(c.get('authorId', 0))
                author = authors[author_id] if author_id < len(authors) else ""
                text = "".join(t.text for t in c.findall(f'.//{ns}t') if t.text).strip()
                if text:
                    comments.append({"type": "cell_comment", "cell": cell_ref, "author": author, "text": text})

    for tf in [f for f in all_files if 'threadedComments' in f and f.endswith('.xml')]:
        xml_bytes = zip_file.read(tf)
        root = ET.fromstring(xml_bytes)
        ns = root.tag.split('}')[0] + '}' if root.tag.startswith('{') else ''
        for tc in root.findall(f'{ns}threadedComment'):
            cell_ref = tc.get('ref', '?')
            text_elem = tc.find(f'{ns}text')
            text = text_elem.text.strip() if text_elem is not None and text_elem.text else ""
            if text:
                comments.append({"type": "threaded_comment", "cell": cell_ref, "text": text,
                                 "is_reply": bool(tc.get('parentId', ''))})

    for vf in [f for f in all_files if 'vmlDrawing' in f and f.endswith('.vml')]:
        try:
            content = zip_file.read(vf).decode('utf-8', errors='replace')
            for dt in re.findall(r'<div[^>]*>(.*?)</div>', content, re.DOTALL):
                clean = re.sub(r'<[^>]+>', '', dt).strip()
                if clean:
                    comments.append({"type": "vml_note", "text": clean})
        except Exception:
            pass

    return comments


# ─── Classification ─────────────────────────────────────

def classify_texts(all_texts):
    assembly_steps = []
    parts = []
    unclassified_notes = []
    current_step = None

    for text in all_texts:
        text = text.strip()
        if not text or text.lower() in VIEW_LABELS:
            continue

        step_match = STEP_PATTERN.match(text)
        if step_match:
            if current_step:
                if unclassified_notes:
                    current_step["cautions"] = list(set(unclassified_notes))
                    unclassified_notes = []
                assembly_steps.append(current_step)
            step_num = int(step_match.group(1))
            instruction = STEP_PATTERN.sub('', text).strip().rstrip('.')
            current_step = {"step": step_num, "instruction": instruction}
            continue

        pn_match = PART_NUMBER_PATTERN.search(text)
        if pn_match:
            lines = text.split('\n')
            part_number = None
            part_name_parts = []
            for line in lines:
                line = line.strip()
                if PART_NUMBER_PATTERN.search(line):
                    part_number = PART_NUMBER_PATTERN.search(line).group()
                else:
                    part_name_parts.append(line)
            if part_number:
                parts.append({"name": " ".join(part_name_parts).strip() or None, "part_number": part_number})
            continue

        is_likely_part = (len(text) < 50
                          and not any(jp in text for jp in ['ください', 'してください', 'ること', 'します', 'ません'])
                          and not text.endswith('。') and not text.endswith('.')
                          and '\n' not in text and not any(c in text for c in ['。', '、']))

        if is_likely_part:
            existing = {p.get("name", "").lower() for p in parts}
            if text.lower() not in existing:
                parts.append({"name": text, "part_number": None})
            continue

        if current_step:
            current_step.setdefault("cautions", [])
            if text not in current_step["cautions"]:
                current_step["cautions"].append(text)
        else:
            if text not in unclassified_notes:
                unclassified_notes.append(text)

    if current_step:
        if unclassified_notes:
            current_step.setdefault("cautions", [])
            current_step["cautions"].extend(unclassified_notes)
        assembly_steps.append(current_step)

    all_instr = " ".join(s.get("instruction", "") + " ".join(s.get("cautions", [])) for s in assembly_steps)
    quoted = re.findall(r'"([^"]+)"', all_instr)
    quoted += re.findall(r'\u201c([^\u201d]+)\u201d', all_instr)
    quoted += re.findall(r'\u300c([^\u300d]+)\u300d', all_instr)
    existing_lower = {p["name"].lower() for p in parts if p.get("name")}
    for qp in quoted:
        qp = qp.strip()
        if qp and qp.lower() not in existing_lower and len(qp) < 60:
            parts.append({"name": qp, "part_number": None})
            existing_lower.add(qp.lower())

    seen = set()
    unique = []
    for p in parts:
        key = (p.get("name", ""), p.get("part_number", ""))
        if key not in seen:
            seen.add(key)
            unique.append(p)

    return assembly_steps, unique


# ─── Image Removal (Zip Surgery) ────────────────────────

def remove_pics_from_element(element, ns_xdr):
    removed = 0
    for pic in element.findall(f'{ns_xdr}pic'):
        element.remove(pic)
        removed += 1
    for grp in element.findall(f'{ns_xdr}grpSp'):
        removed += remove_pics_from_element(grp, ns_xdr)
    return removed


def remove_images_from_zip(input_path, output_path):
    with zipfile.ZipFile(input_path, 'r') as zin:
        all_files = zin.namelist()
        image_files = [f for f in all_files if f.startswith('xl/media/')]
        drawing_files = [f for f in all_files if f.startswith('xl/drawings/') and f.endswith('.xml') and '_rels' not in f]

        modified_drawings = {}
        total_removed = 0

        for df in drawing_files:
            drawing_xml = zin.read(df).decode("utf-8")
            for prefix, uri in re.findall(r'xmlns:?(\w*)="([^"]+)"', drawing_xml):
                if prefix:
                    try:
                        ET.register_namespace(prefix, uri)
                    except ValueError:
                        pass

            root = ET.fromstring(drawing_xml)
            ns_xdr = '{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}'
            ns_a = '{http://schemas.openxmlformats.org/drawingml/2006/main}'

            anchors_to_remove = []
            for anchor in list(root):
                if 'Anchor' not in anchor.tag:
                    continue
                has_pic = anchor.find(f'.//{ns_xdr}pic') is not None
                has_shape = anchor.find(f'.//{ns_xdr}sp') is not None
                has_group = anchor.find(f'.//{ns_xdr}grpSp') is not None
                has_connector = anchor.find(f'.//{ns_xdr}cxnSp') is not None
                texts = [t.text for t in anchor.findall(f'.//{ns_a}t') if t.text]

                if has_pic and not has_shape and not has_group and not has_connector and not texts:
                    anchors_to_remove.append(anchor)
                    total_removed += len(anchor.findall(f'.//{ns_xdr}pic'))
                elif has_pic:
                    total_removed += remove_pics_from_element(anchor, ns_xdr)

            for a in anchors_to_remove:
                root.remove(a)

            buf = io.BytesIO()
            ET.ElementTree(root).write(buf, xml_declaration=True, encoding='UTF-8')
            modified_drawings[df] = buf.getvalue()

        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in all_files:
                if item in image_files:
                    continue
                elif item in modified_drawings:
                    zout.writestr(item, modified_drawings[item])
                else:
                    zout.writestr(item, zin.read(item))

    return total_removed


# ─── Write Extracted Data Sheet ──────────────────────────

def write_data_sheet(filepath, assembly_steps, parts, comments):
    wb = load_workbook(filepath)

    if EXTRACTED_SHEET_NAME in wb.sheetnames:
        del wb[EXTRACTED_SHEET_NAME]

    ws = wb.create_sheet(EXTRACTED_SHEET_NAME)

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    section_font = Font(bold=True, size=11, color="1F4E79")
    section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    step_font = Font(bold=True, size=10)
    caution_font = Font(italic=True, size=10, color="CC0000")
    normal_font = Font(size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    row = 1

    # ── ASSEMBLY STEPS ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value="ASSEMBLY STEPS")
    cell.font = header_font
    cell.fill = header_fill
    for c in range(1, 4):
        ws.cell(row=row, column=c).fill = header_fill
    row += 1

    for i, h in enumerate(["Step", "Instruction", "Cautions / Notes"], 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = section_font
        cell.fill = section_fill
    row += 1

    if assembly_steps:
        for step in assembly_steps:
            ws.cell(row=row, column=1, value=step["step"]).font = step_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row=row, column=2, value=step["instruction"]).font = normal_font
            ws.cell(row=row, column=2).alignment = wrap_align
            cautions = step.get("cautions", [])
            if cautions:
                ws.cell(row=row, column=3, value="\n".join(f"⚠ {c}" for c in cautions)).font = caution_font
                ws.cell(row=row, column=3).alignment = wrap_align
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = thin_border
            row += 1
    else:
        ws.cell(row=row, column=1, value="(No assembly steps found)").font = Font(italic=True, color="888888")
        row += 1

    row += 1

    # ── PARTS ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value="PARTS REFERENCED")
    cell.font = header_font
    cell.fill = header_fill
    for c in range(1, 4):
        ws.cell(row=row, column=c).fill = header_fill
    row += 1

    for i, h in enumerate(["#", "Part Name", "Part Number"], 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = section_font
        cell.fill = section_fill
    row += 1

    if parts:
        for idx, part in enumerate(parts, 1):
            ws.cell(row=row, column=1, value=idx).font = normal_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=part.get("name", "")).font = normal_font
            ws.cell(row=row, column=3, value=part.get("part_number", "") or "—").font = normal_font
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = thin_border
            row += 1
    else:
        ws.cell(row=row, column=1, value="(No parts found)").font = Font(italic=True, color="888888")
        row += 1

    row += 1

    # ── COMMENTS (if any) ──
    if comments:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value="COMMENTS")
        cell.font = header_font
        cell.fill = header_fill
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = header_fill
        row += 1

        for i, h in enumerate(["Cell", "Author", "Comment"], 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = section_font
            cell.fill = section_fill
        row += 1

        for c in comments:
            ws.cell(row=row, column=1, value=c.get("cell", "")).font = normal_font
            ws.cell(row=row, column=2, value=c.get("author", "")).font = normal_font
            ws.cell(row=row, column=3, value=c.get("text", "")).font = normal_font
            ws.cell(row=row, column=3).alignment = wrap_align
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 50

    wb.save(filepath)


# ─── Main ────────────────────────────────────────────────

def process_excel(input_path, output_path):
    filename = os.path.basename(input_path)
    try:
        with zipfile.ZipFile(input_path, 'r') as z:
            all_texts = extract_texts_from_zip(z)
            comments = extract_all_comments(z)
            image_count = len([f for f in z.namelist() if f.startswith('xl/media/')])

        assembly_steps, parts = classify_texts(all_texts) if all_texts else ([], [])

        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)
        try:
            images_removed = remove_images_from_zip(input_path, temp_path)
            write_data_sheet(temp_path, assembly_steps, parts, comments)
            shutil.copy2(temp_path, output_path)
        finally:
            os.unlink(temp_path)

        in_size = os.path.getsize(input_path) / (1024 * 1024)
        out_size = os.path.getsize(output_path) / (1024 * 1024)
        print(f"    → {len(assembly_steps)} steps, {len(parts)} parts, {len(comments)} comments")
        print(f"    → {images_removed} images removed  ({in_size:.2f} MB → {out_size:.2f} MB)")
        return True

    except zipfile.BadZipFile:
        print(f"    ✗ Not a valid xlsx file")
        return False
    except Exception as e:
        print(f"    ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    base_dir = get_base_dir()
    print(f"Working directory: {base_dir}")

    output_dir = os.path.join(base_dir, OUTPUT_FOLDER_NAME)
    output_dir_normalized = os.path.normpath(output_dir)

    excel_files = []
    for f in sorted(glob.glob(os.path.join(base_dir, "*.xlsx"))):
        if os.path.normpath(os.path.dirname(f)) == output_dir_normalized:
            continue
        excel_files.append(f)

    if not excel_files:
        print(f"\nNo Excel files (.xlsx) found in: {base_dir}")
        print("Make sure .xlsx files are in the same folder, or run from the folder containing them.")
        exit(1)

    print(f"Found {len(excel_files)} Excel file(s)\n")
    os.makedirs(output_dir, exist_ok=True)

    success = 0
    failed = 0

    for filepath in excel_files:
        filename = os.path.basename(filepath)
        output_path = os.path.join(output_dir, filename)
        print(f"  Processing: {filename}")

        if process_excel(filepath, output_path):
            print(f"    → Saved: {OUTPUT_FOLDER_NAME}/{filename}")
            success += 1
        else:
            failed += 1

    print(f"\n{'='*50}")
    print(f"Done! {success} processed, {failed} failed.")
    print(f'Output saved to: {output_dir}')
    print(f'\nOpen any file → look for the "{EXTRACTED_SHEET_NAME}" sheet tab.')


if __name__ == "__main__":
    main()

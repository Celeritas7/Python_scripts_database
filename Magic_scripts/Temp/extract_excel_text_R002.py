"""
Remove Images from ALL Excel Files (Keep Everything Else)
==========================================================
Usage:
    pip install openpyxl
    python extract_excel_text.py

This script preserves ALL formatting, borders, colors, merged cells,
and text — it ONLY removes embedded images and charts.

Cleaned files are saved in a 'cleaned' subfolder.
"""

from openpyxl import load_workbook
import os
import glob


def get_base_dir():
    """
    Smart directory detection:
    - If called from a different folder (e.g. via tools, IDE, or another script),
      use the current working directory.
    - If run directly from its own folder, use the script's directory.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    work_dir = os.getcwd()

    if work_dir != script_dir:
        base_dir = work_dir
    else:
        base_dir = script_dir

    return base_dir


def remove_images(input_path, output_path):
    print(f"\n  Loading: {os.path.basename(input_path)} ...")

    # Load the full workbook (preserves all formatting, styles, merges, etc.)
    wb = load_workbook(input_path)

    total_removed = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        img_count = len(ws._images)
        chart_count = len(ws._charts)

        if img_count > 0 or chart_count > 0:
            print(f"    Sheet '{sheet_name}': removing {img_count} image(s), {chart_count} chart(s)")
            ws._images = []
            ws._charts = []
            total_removed += img_count + chart_count
        else:
            print(f"    Sheet '{sheet_name}': no images/charts found")

    wb.save(output_path)
    input_size = os.path.getsize(input_path) / (1024 * 1024)
    output_size = os.path.getsize(output_path) / (1024 * 1024)
    print(f"    {input_size:.1f} MB --> {output_size:.1f} MB  ({total_removed} removed)  ✓")


if __name__ == "__main__":
    base_dir = get_base_dir()
    print(f"Working directory: {base_dir}")

    # Find all Excel files in the folder
    excel_files = sorted(
        glob.glob(os.path.join(base_dir, "*.xlsx")) +
        glob.glob(os.path.join(base_dir, "*.xls"))
    )

    if not excel_files:
        print("\nNo Excel files (.xlsx / .xls) found in this folder.")
        print("Please run this script from the folder containing your Excel files.")
        exit(1)

    print(f"Found {len(excel_files)} Excel file(s)")

    # Create 'cleaned' subfolder for output
    output_dir = os.path.join(base_dir, "cleaned")
    os.makedirs(output_dir, exist_ok=True)

    success = 0
    failed = 0

    for filepath in excel_files:
        filename = os.path.basename(filepath)
        output_path = os.path.join(output_dir, filename)

        try:
            remove_images(filepath, output_path)
            success += 1
        except Exception as e:
            print(f"\n  ✗ Failed: {filename} — {e}")
            failed += 1

    print(f"\n{'='*50}")
    print(f"Complete! {success} processed, {failed} failed.")
    print(f"Cleaned files saved to: {output_dir}")

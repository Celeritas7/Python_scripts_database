"""
Extract Text Data from Excel (Remove Images & Embedded Objects)
===============================================================
Usage:
    pip install openpyxl
    python extract_excel_text.py

Place this script in the same folder as your Excel file, or call it
from the folder containing your Excel file.

Update INPUT_FILENAME and OUTPUT_FILENAME below before running.
"""

from openpyxl import load_workbook
from openpyxl import Workbook
import os

# ========== CONFIGURE THESE ==========
INPUT_FILENAME = "your_file.xlsx"       # <-- Change to your input filename
OUTPUT_FILENAME = "cleaned_output.xlsx"  # <-- Change to your desired output filename
# ======================================


def get_base_dir():
    """
    Smart directory detection:
    - If called from a different folder (e.g. via tools, IDE, or another script),
      use the current working directory.
    - If run directly from its own folder, use the script's directory.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    work_dir = os.getcwd()

    if work_dir != script_dir:
        # Called from a different folder — use that folder
        base_dir = work_dir
    else:
        # Run directly from its own folder — use script's folder
        base_dir = script_dir

    return base_dir


def extract_text_only(input_path, output_path):
    print(f"Loading: {input_path} ...")
    # Load workbook as data-only (resolves formulas to values)
    wb_source = load_workbook(input_path, data_only=True)
    wb_new = Workbook()
    wb_new.remove(wb_new.active)  # Remove default sheet

    for sheet_name in wb_source.sheetnames:
        ws_source = wb_source[sheet_name]
        ws_new = wb_new.create_sheet(title=sheet_name)

        print(f"  Processing sheet: '{sheet_name}' ({ws_source.max_row} rows x {ws_source.max_column} cols)")

        # Copy cell values only (no images, charts, or objects)
        for row in ws_source.iter_rows(min_row=1, max_row=ws_source.max_row,
                                        min_col=1, max_col=ws_source.max_column):
            for cell in row:
                new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)

                # Preserve basic formatting (optional — remove this block if not needed)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.alignment = cell.alignment.copy()
                    new_cell.number_format = cell.number_format

        # Copy column widths
        for col_letter, dim in ws_source.column_dimensions.items():
            ws_new.column_dimensions[col_letter].width = dim.width

        # Copy merged cells
        for merged_range in ws_source.merged_cells.ranges:
            ws_new.merge_cells(str(merged_range))

    wb_new.save(output_path)
    input_size = os.path.getsize(input_path) / (1024 * 1024)
    output_size = os.path.getsize(output_path) / (1024 * 1024)
    print(f"\nDone!")
    print(f"  Input:  {input_size:.1f} MB")
    print(f"  Output: {output_size:.1f} MB")
    print(f"  Saved to: {output_path}")


if __name__ == "__main__":
    base_dir = get_base_dir()
    print(f"Working directory: {base_dir}")

    input_path = os.path.join(base_dir, INPUT_FILENAME)
    output_path = os.path.join(base_dir, OUTPUT_FILENAME)

    if not os.path.exists(input_path):
        print(f"Error: File '{input_path}' not found.")
        print("Please update INPUT_FILENAME in the script or run it from the folder containing your Excel file.")
    else:
        extract_text_only(input_path, output_path)
